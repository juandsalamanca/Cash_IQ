from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def style_projections(OUTPUT_XLSX, inflow_section_indexes, outflow_section_indexes, cash_balance_indexes):

    wb = load_workbook(OUTPUT_XLSX)

    # Define styles
    sheet_name = "Projections (Table)"

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]
    # Styles
    font_style = Font(name="Aptos Narrow", size=12)
    header_fill = PatternFill(
        start_color="A3A5D0",
        end_color="A3A5D0",
        fill_type="solid"
    )

    accounting_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

    # Apply styles
    for row in ws.iter_rows():
        for cell in row:
            # Font for everything
            cell.font = font_style

            # Accounting format only for numbers
            if isinstance(cell.value, (int, float)):
                cell.number_format = accounting_format

    # Header fill
    for cell in ws[1]:
        cell.fill = header_fill

    # Category fill
    category_fill = PatternFill(
        start_color="BFBFBF",
        end_color="BFBFBF",
        fill_type="solid"
    )

    for section_indexes in [inflow_section_indexes, outflow_section_indexes]:
        
        for i in range(len(section_indexes)):
            idx = section_indexes[i]
            row = ws[idx]

            # Total inflows and outflows are one column to the left of the categories
            if i == len(section_indexes)-1:
                start = 0
            else:
                start = 1
            for i in range(start, len(row)):
                row[i].fill = category_fill

    # Apply color to bag end cash
    beg_cash_row_idx = cash_balance_indexes[0]
    end_cash_row_idx = cash_balance_indexes[1]
    beg_row = ws[beg_cash_row_idx]
    end_row = ws[end_cash_row_idx]

    for col in range(ws.max_column):
        beg_row[col].fill = header_fill
        end_row[col].fill = header_fill


    wb.save(OUTPUT_XLSX)