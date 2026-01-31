import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from src.trinity.classify_transactions import classify_inflows, classify_outflows

def get_combined_bank(proj_bank, bank_actual_pivot, actual_week_starts, proj_week_starts, all_week_starts, cc_payment_alloc):
    # Add CC payment allocation rows to bank cash projections
    if len(cc_payment_alloc):
        # ensure columns match
        cc_payment_alloc = cc_payment_alloc.reindex(columns=proj_week_starts, fill_value=0.0)
        # append to proj_bank (cash impacts)
        for idx in cc_payment_alloc.index:
            if idx not in proj_bank.index:
                proj_bank.loc[idx] = 0.0
            proj_bank.loc[idx, proj_week_starts] += cc_payment_alloc.loc[idx, proj_week_starts].values

    # =========================
    # COMBINE ACTUALS + PROJECTIONS FOR BANK CASH
    # =========================
    combined_bank = pd.concat(
        [
            bank_actual_pivot[actual_week_starts],
            proj_bank[proj_week_starts]
        ],
        axis=1
    ).fillna(0.0)

    combined_full = combined_bank.reindex(columns=all_week_starts, fill_value=0.0)
    return combined_full

def collapse_other(df, keep_index, other_name, index_names):
    keep = df.loc[keep_index].copy() if len(keep_index) else df.iloc[0:0].copy()
    other = df.drop(index=keep_index, errors="ignore").copy()
    if len(other):
        other_row = other.sum(axis=0)
        other_idx = pd.MultiIndex.from_tuples([(other_name, "Other", "")], names=index_names)
        other_df = pd.DataFrame([other_row.values], index=other_idx, columns=df.columns)
        keep = pd.concat([keep, other_df], axis=0)
    return keep

def build_inflows_outflows(combined_full, actual_week_starts, all_week_starts, TOP_N_INFLOW_LINES, TOP_N_OUTFLOW_LINES, idx_names):
    # =========================
    # BUILD INFLOWS/OUTFLOWS PRESENTATION (NO FLAT)
    # =========================
    trailing_actual = combined_full[actual_week_starts].copy()
    inflow_mask  = trailing_actual.sum(axis=1) > 0
    outflow_mask = trailing_actual.sum(axis=1) < 0

    # rank lines by trailing magnitude
    top_inflows = (
        combined_full.loc[inflow_mask]
        .assign(trailing_inflow=lambda df: trailing_actual.loc[df.index].clip(lower=0).sum(axis=1))
        .sort_values("trailing_inflow", ascending=False)
        .head(TOP_N_INFLOW_LINES)
        .index
    )

    top_outflows = (
        combined_full.loc[outflow_mask]
        .assign(trailing_outflow=lambda df: (-trailing_actual.loc[df.index].clip(upper=0)).sum(axis=1))
        .sort_values("trailing_outflow", ascending=False)
        .head(TOP_N_OUTFLOW_LINES)
        .index
    )

    inflows_tbl = combined_full.loc[inflow_mask, all_week_starts].copy()
    outflows_tbl = combined_full.loc[outflow_mask, all_week_starts].copy()

    inflows_tbl  = collapse_other(inflows_tbl,  top_inflows,  "Other Inflows",  idx_names)
    outflows_tbl = collapse_other(outflows_tbl, top_outflows, "Other Outflows", idx_names)

    # Presentation: inflows positive; outflows positive
    inflows_present  = inflows_tbl.copy()
    outflows_present = outflows_tbl.copy().abs()

    total_inflows  = inflows_present.sum(axis=0)
    total_outflows = outflows_present.sum(axis=0)

    return inflows_present, outflows_present, total_inflows, total_outflows

def get_cash_balance(total_inflows, total_outflows, beginning_cash_balance, all_week_starts):
    # =========================
    # BEGIN/END CASH BALANCE
    # =========================
    beg_bal_series = pd.Series(index=all_week_starts, dtype=float)
    end_bal_series = pd.Series(index=all_week_starts, dtype=float)

    running_begin = beginning_cash_balance
    for w in all_week_starts:
        beg_bal_series[w] = running_begin
        running_end = running_begin + float(total_inflows[w]) - float(total_outflows[w])
        end_bal_series[w] = running_end
        running_begin = running_end

    return beg_bal_series, end_bal_series

    

def get_cc_output_sheets(cc_spend_cat_pivot_top, cc_spend_proj_cat, cc_payment_alloc, all_week_starts, proj_week_starts):
    # =========================
    # CC OUTPUT SHEETS:
    #   - CC Spend Transactions (all CC spend rows)
    #   - CC Spend Weekly (by category)
    #   - CC Spend Projected (by category)
    #   - CC Payment Schedule
    #   - CC Payment Allocation (cash impact, by category)
    # =========================
    # Weekly CC spend actual (by category) for display (last actual + projected horizon)
    cc_spend_actual_display = cc_spend_cat_pivot_top.reindex(columns=all_week_starts, fill_value=0.0)
    for w in cc_spend_actual_display.columns:
        if w not in cc_spend_cat_pivot_top.columns:
            cc_spend_actual_display[w] = 0.0
    cc_spend_actual_display = cc_spend_actual_display[all_week_starts]

    cc_spend_proj_display = cc_spend_proj_cat.reindex(columns=proj_week_starts, fill_value=0.0)

    cc_payment_alloc_present = cc_payment_alloc.abs() if len(cc_payment_alloc) else pd.DataFrame(columns=proj_week_starts)

    return cc_spend_proj_display, cc_spend_actual_display, cc_payment_alloc_present



def write_output_excel(all_week_starts, inflows_by_cat, outflows_by_cat, inflows_present, outflows_present, total_inflows, total_outflows, cc_spend_proj_display, cc_spend_actual_display, cc_payment_alloc_present, cc_spend_txn, cc_payment_schedule, beg_bal_series, end_bal_series, PROJ_WEEK1_START, OUTPUT_XLSX):
    # =========================
    # WRITE OUTPUT EXCEL
    # =========================
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        # Summary
        summary = pd.DataFrame(
            {
                "Week Start": all_week_starts,
                "Beginning Bank Balance": [beg_bal_series[w] for w in all_week_starts],
                "Total Cash Inflows": [total_inflows[w] for w in all_week_starts],
                "Total Cash Outflows": [total_outflows[w] for w in all_week_starts],
                "Ending Bank Balance": [end_bal_series[w] for w in all_week_starts],
            }
        )
        summary.to_excel(writer, sheet_name="Summary", index=False)

        # Cash details
        inflows_present.reset_index().to_excel(writer, sheet_name="Cash Inflows (Detail)", index=False)
        outflows_present.reset_index().to_excel(writer, sheet_name="Cash Outflows (Detail)", index=False)

        # Credit card sheets
        cc_spend_txn.sort_values(["account_name","date"]).to_excel(writer, sheet_name="CC Spend - Transactions", index=False)
        cc_spend_actual_display.reset_index().to_excel(writer, sheet_name="CC Spend - Weekly (Hist)", index=False)
        cc_spend_proj_display.reset_index().to_excel(writer, sheet_name="CC Spend - Weekly (Proj)", index=False)
        cc_payment_schedule.to_excel(writer, sheet_name="CC Payments - Schedule", index=False)
        cc_payment_alloc_present.reset_index().to_excel(writer, sheet_name="Cash - CC Pay Allocation", index=False)

        # Template-style table
        rows = []
        rows.append(("Beginning Bank Balance", "", ""))
        rows.append(("Cash Inflows", "", ""))
        inflow_section_indexes = []
        for inflow_cat in inflows_by_cat:
            rows.append(("", inflow_cat, ""))
            inflow_section_indexes.append(len(rows)+1)
            for acct in sorted(inflows_by_cat[inflow_cat]):
                rows.append(("", "", acct))

            rows.append(("", "", ""))

        rows.append(("Total Cash Inflows", "", ""))
        inflow_section_indexes.append(len(rows)+1)
        rows.append(("Cash Outflows", "", ""))
        outflow_section_indexes = []
        for outflow_cat in outflows_by_cat:
            rows.append(("", outflow_cat, ""))
            outflow_section_indexes.append(len(rows)+1)
            for acct in sorted(outflows_by_cat[outflow_cat]):
                rows.append(("", "", acct))

            rows.append(("", "", ""))

        rows.append(("Total Cash Outflows", "", ""))
        outflow_section_indexes.append(len(rows)+1)
        rows.append(("Ending Bank Balance", "", ""))

        proj_sheet = pd.DataFrame(rows, columns=["Section","Notes","Line Item"])
        for w in all_week_starts:
            proj_sheet[w.strftime("%Y-%m-%d")] = np.nan

        def put_row_value(section, acct, values):
            mask = (proj_sheet["Section"].eq(section)) & (proj_sheet["Line Item"].eq(acct))
            idx = proj_sheet.index[mask]
            if len(idx):
                i = idx[0]
                for w in all_week_starts:
                    proj_sheet.loc[i, w.strftime("%Y-%m-%d")] = float(values[w])

        put_row_value("Beginning Bank Balance","", beg_bal_series)
        put_row_value("Ending Bank Balance","", end_bal_series)
        put_row_value("Total Cash Inflows","", total_inflows)
        put_row_value("Total Cash Outflows","", total_outflows)

        for (acct, typ, det), row in inflows_present.iterrows():
            put_row_value("", acct, row)

        for (acct, typ, det), row in outflows_present.iterrows():
            put_row_value("", acct, row)

        # Switch the Notes and Line item columns for the projections sheet
        #proj_sheet[['Line Item', 'Notes']] = proj_sheet[['Notes', 'Line Item']].values
        #proj_sheet = proj_sheet.rename(columns={'Line Item':'Notes','Notes':'Line Item'})

        proj_sheet.to_excel(writer, sheet_name="Projections (Table)", index=False)

        print(f"Saved: {OUTPUT_XLSX}")
        print(f"Projection Week 1 starts: {PROJ_WEEK1_START.date()} (Monday)")

    return inflow_section_indexes, outflow_section_indexes


def calculate_category_totals(OUTPUT_XLSX, inflow_section_indexes, outflow_section_indexes):

    wb = load_workbook(OUTPUT_XLSX, data_only=True)
    ws = wb["Projections (Table)"]

    for section_indexes in [inflow_section_indexes, outflow_section_indexes]:
    
        for i in range(len(section_indexes)):
            idx = section_indexes[i]
            if idx == section_indexes[-1]:
                break
            
            next_idx = section_indexes[i+1]
            
            row = ws[idx]
            for col in range(3, len(row)):
                col_letter = get_column_letter(col+1)
                if next_idx-2 >= idx+1:
                    row[col].value = f'=SUM({col_letter}{idx+1}:{col_letter}{next_idx-2})'
                else:
                    row[col].value = 0.0

    wb.save(OUTPUT_XLSX)